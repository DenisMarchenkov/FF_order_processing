from datetime import datetime
from pathlib import Path

import pandas
import openpyxl
from tabulate import tabulate

from services import *

SETTINGS = get_settings()
HEADERS = ['НОМЕР ЗАКАЗА', 'ДАТА ЗАКАЗА', 'МАРКА', 'ВН. КОД', 'АРТИКУЛ', 'НАИМЕНОВАНИЕ',
           'ШТ', 'СРОК ГОДНОСТИ', 'ЦЕНА С НДС', 'ИТОГО С НДС', 'СУММА ЗАКАЗА']

HEADERS_FF = ['ВН. КОД', 'НАИМЕНОВАНИЕ', 'ШТ', 'НОМЕР ЗАКАЗА']

save_attachment_all_email(
    SETTINGS["server_imap"],
    SETTINGS["login"],
    SETTINGS["password_api"],
    SETTINGS["download_dir"],
    SETTINGS["download_dir_for_file_FF"],
)

files = Path(SETTINGS["download_dir"]).glob('*.xls')
frames = [get_dataframe(file) for file in files]

# files_FF = Path(SETTINGS["download_dir_for_file_FF"]).glob('*.csv')
# frames_FF = [get_dataframe_csv(file) for file in files_FF]
# common_frame_FF = pandas.concat(frames_FF)
# common_frame_FF.reset_index(inplace=True, drop=True)
# common_frame_FF.columns = HEADERS_FF
# #common_frame_FF['ВН. КОД'].astype(int)
# print(common_frame_FF)

print(f'Не обработанных файлов: {len(frames)}')
if len(frames) != 0:
    id_recap = datetime.now().strftime("%d%m%y-%H%M%S")
    name_recap = f'recap-{id_recap}.xlsx'
    wb = openpyxl.Workbook()
    wb.save(Path(SETTINGS["cur_dir"], name_recap))

    dataframe = pandas.concat(frames)
    dataframe.columns = HEADERS
    dataframe['ВН. КОД'] = dataframe['ВН. КОД'].astype(int)
    # print(dataframe)
    #
    # commons = pandas.merge(common_frame_FF, dataframe, how='inner', on=['ВН. КОД'])['ВН. КОД']
    # print('ОБЩЕЕ')
    # print(commons)
    # differences = common_frame_FF[~common_frame_FF['ВН. КОД'].isin(commons)]
    # print('Различия')
    # print(differences)

    recap = pandas.pivot_table(dataframe, values=["ШТ", 'ИТОГО С НДС'],
                               index=['НОМЕР ЗАКАЗА', 'ДАТА ЗАКАЗА'],
                               aggfunc="sum")

    table = pandas.pivot_table(dataframe, values=["ШТ"],
                               index=['МАРКА', 'АРТИКУЛ', 'НАИМЕНОВАНИЕ', 'СРОК ГОДНОСТИ'],
                               aggfunc="sum",)

    with pandas.ExcelWriter(name_recap, mode="a", if_sheet_exists='overlay') as writer:
        recap.to_excel(writer, sheet_name="Recap", startrow=2)
        table.to_excel(writer, sheet_name="Сводная таблица товаров", startrow=2)
        #print(table.size)
        #differences.to_excel(writer, sheet_name="Сводная таблица товаров", startrow=table.size + 3, startcol=1,index=False, header=False)

        for frame in frames:
            sheet_name = f'Order {frame[0].loc[frame.index[0]]}'
            frame.columns = HEADERS
            frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)



    wb = openpyxl.load_workbook(Path(SETTINGS["cur_dir"], name_recap))
    del wb["Sheet"]
    sheet_recap = wb["Recap"]
    sheet_pivot_table = wb["Сводная таблица товаров"]
    sheets_orders = wb.sheetnames[2:]

    last_pivot = last_row_col(sheet_pivot_table)
    last_recap = last_row_col(sheet_recap)

    sheet_format(sheet=sheet_recap, column_date=[2], **last_recap, recap=True, file_name=name_recap)
    sheet_format(sheet=sheet_pivot_table, column_date=[4], **last_pivot, recap=False, file_name=name_recap)

    for sheet in sheets_orders:
        sheets_orders_formatting(sheet=wb[sheet], column_date=[2, 8], **last_row_col(wb[sheet]), file_name=name_recap)

    wb.save(Path(SETTINGS["cur_dir"], name_recap))
    movement_files(SETTINGS["download_dir"], SETTINGS["completed_dir"], '.xls')

    text = tabulate(text_message(recap),
                    headers='keys',
                    tablefmt='html',
                    floatfmt=(None, ".2f", None, None),
                    colalign=(None, "decimal", "center", "right",))
    html = text.replace('table', 'table border="1"; style="border-collapse:collapse; width: 300px;"',)

    send_email(
        SETTINGS['server_smtp'],
        SETTINGS['login'],
        SETTINGS['password_api'],
        Path(SETTINGS['cur_dir'], name_recap),
        message=html,
        file_name=name_recap,
    )
    movement_files(SETTINGS["cur_dir"], SETTINGS["recap_dir"], '.xlsx')
print('Обработка завершена')

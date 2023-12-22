def get_settings():
    with open('settings.txt', 'r') as fl:
        dt = fl.readlines()
        dt = [line.rstrip() for line in dt]

        from pathlib import Path
        cur_dir = Path.cwd()
        download_dir = Path(cur_dir, 'Download')
        completed_dir = Path(download_dir, 'Completed')
        server_imap = dt[0].split('= ')[1]
        port_imap = dt[1].split('= ')[1]
        server_smtp = dt[2].split('= ')[1]
        port_smtp = dt[3].split('= ')[1]
        login = dt[4].split('= ')[1]
        password_api = dt[5].split('= ')[1]

        settings = {
            "cur_dir": cur_dir,
            "download_dir": download_dir,
            "completed_dir": completed_dir,
            "server_imap": server_imap,
            "port_imap": port_imap,
            "server_smtp": server_smtp,
            "port_smtp": port_smtp,
            "login": login,
            "password_api": password_api,
        }
        return settings


def save_attachment_all_email(server, login, passw, save_path):
    import imaplib
    import email
    import os

    print("-- подключаемся к", server)
    mail = imaplib.IMAP4_SSL(server)
    print("-- логинимся")
    mail.login(login, passw)
    mail.list()
    print("-- подключаемся к inbox")
    mail.select("inbox")
    print("-- получаем UID последнего письма")
    result, data = mail.uid('search', None, "ALL")

    for num in data[0].split():
        result, data = mail.uid('fetch', num, '(RFC822)')
        raw_email = data[0][1]
        try:
            email_message = email.message_from_string(raw_email)
        except TypeError:
            email_message = email.message_from_bytes(raw_email)
        print("нашли письмо от: ", email.header.make_header(email.header.decode_header(email_message['From'])))
        for part in email_message.walk():
            if "application" in part.get_content_type():
                filename = part.get_filename()
                filename = str(email.header.make_header(email.header.decode_header(filename)))
                file_extension = filename.split('.')[1]
                if file_extension == 'xls':
                    print("нашли вложение ", filename)
                    fp = open(os.path.join(save_path, filename), 'wb')
                    fp.write(part.get_payload(decode=1))
                    fp.close
                    print(f'файл {filename} сохранен')
        print("-- перемещаем письмо в папку Completed")
        copy_res = mail.uid('COPY', num, 'INBOX/Completed')
        if copy_res[0] == 'OK':
            delete_res = mail.uid('STORE', num, '+FLAGS', '(\Deleted)')
        # print('*************')
        mail.expunge()
    mail.close()
    mail.logout()


def get_dataframe(path):
    from pandas import read_excel
    dataframe = read_excel(path, header=None, na_values='не указан')
    dataframe = dataframe[1:]
    dataframe = dataframe.fillna('нет данных')
    return dataframe


def sheet_format(sheet, last_empty_row, last_empty_columns, column_date, recap, file_name):
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.styles.numbers import BUILTIN_FORMATS


    thin_border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'),
    )
    alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
    )

    for row in range(3, last_empty_row):
        if recap:
            sheet.cell(row, 3).number_format = BUILTIN_FORMATS[4]

        for column in range(1, last_empty_columns):
            cell = sheet.cell(row, column)
            if row == 3:
                cell.alignment = alignment
                cell.font = Font(bold=True)
            else:
                for item in column_date:
                    sheet.cell(row, item).number_format = 'DD.MM.YYYY'
                cell.font = Font(name='Calibri', size=14)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = thin_border
                sheet.cell(row, last_empty_columns - 1).alignment = alignment

    cm = 1 / 2.54
    sheet.page_margins = PageMargins(left=cm * 0.8,
                                     right=cm * 0.8,
                                     top=cm * 0.8,
                                     bottom=cm * 1.8)

    if recap:
        sheet.cell(1, 3).font = Font(name='Calibri', size=20)
        sheet.cell(1, 3).value = file_name[:-5]
        sheet.column_dimensions['A'].width = 11
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 5
        sheet.oddFooter.left.text = file_name[:-5]
        sheet.print_title_rows = "3:3"
    else:
        sheet.cell(1, 1).font = Font(name='Calibri', size=20)
        sheet.cell(1, 1).value = f'СВОДНАЯ ТАБЛИЦА ТОВАРОВ К {file_name[:-5]}'
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 17
        sheet.column_dimensions['C'].width = 44
        sheet.column_dimensions['D'].width = 14
        sheet.column_dimensions['E'].width = 5
        sheet.oddFooter.left.text = file_name[:-5]
        sheet.print_title_rows = "3:3"
        for i in range(4, last_empty_row):
            cell = sheet.cell(i, 5)
            if cell.value > 1:
                cell.fill = PatternFill(fill_type='solid', fgColor='FF000000')
                cell.font = Font(color='ffffff', name='Calibri', size=14)
                cell.border = Border(bottom=Side(border_style='thin', color='ffffff'))


def sheets_orders_formatting(sheet, last_empty_row, last_empty_columns, column_date, file_name):
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.styles.numbers import BUILTIN_FORMATS

    thin_border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000'),
                         )
    total_qua = []
    for i in range(3, last_empty_row):
        for j in range(1, last_empty_columns):
            cell = sheet.cell(i, j)
            for item in column_date:
                sheet.cell(i, item).number_format = 'DD.MM.YYYY'
            cell.font = Font(name='Calibri', size=14)
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = thin_border
            if j == 7:  # column qua
                total_qua.append(cell.value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if j == 9:  # column price
                cell.number_format = BUILTIN_FORMATS[4]
            if i == 3:  # header
                sheet.cell(3, j).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    for i in range(4, last_empty_row):
        cell = sheet.cell(i, 7)
        if cell.value > 1:
            cell.fill = PatternFill(fill_type='solid', fgColor='FF000000')
            cell.font = Font(color='ffffff', name='Calibri', size=14)
            cell.border = Border(bottom=Side(border_style='thin', color='ffffff'))

    total_qua = total_qua[1:]

    sheet.insert_rows(2, amount=4)
    sheet.cell(1, 3).value = "НОМЕР ЗАКАЗА"
    sheet.cell(2, 3).value = "ДАТА ЗАКАЗА"
    sheet.cell(3, 3).value = "СУММА ЗАКАЗА"
    sheet.cell(4, 3).value = "СТРОК В ЗАКАЗЕ"
    sheet.cell(5, 3).value = "ШТУК В ЗАКАЗЕ"
    sheet.cell(1, 6).value = int(sheet.cell(8, 1).value)
    sheet.cell(2, 6).value = sheet.cell(8, 2).value
    sheet.cell(2, 6).number_format = 'DD.MM.YYYY'
    sheet.cell(3, 6).value = sheet.cell(8, 11).value
    sheet.cell(3, 6).number_format = BUILTIN_FORMATS[4]
    sheet.cell(4, 6).value = len(total_qua)
    sheet.cell(5, 6).value = sum(total_qua)

    sheet.delete_cols(1, 2)
    sheet.delete_cols(2)
    sheet.delete_cols(7, 2)

    row = 1
    while row <= 5:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        col = 1
        while col <= 3:
            sheet.cell(row, col).alignment = Alignment(horizontal='center')
            sheet.cell(row, col).border = thin_border
            sheet.cell(row, col).font = Font(name='Calibri', size=16)
            col += 1
        row += 1

    sheet.column_dimensions['A'].width = 14
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 36
    sheet.column_dimensions['D'].width = 4
    sheet.column_dimensions['E'].width = 14
    sheet.column_dimensions['F'].width = 13

    cm = 1 / 2.54
    sheet.page_margins = PageMargins(left=cm * 0.8,
                                     right=cm * 0.8,
                                     top=cm * 0.8,
                                     bottom=cm * 1.8)
    sheet.print_options.horizontalCentered = True
    sheet.oddFooter.right.text = f'Заказ {sheet.cell(1, 3).value}' # + "страница &[Page] из &N"
    sheet.oddFooter.left.text = file_name[:-5]
    sheet.print_title_rows = "7:7"


def last_row_col(sheet):
    last = {
        "last_empty_row": sheet.max_row + 1,
        "last_empty_columns": sheet.max_column + 1
    }
    return last


def movement_files(file_source, file_destination, file_extension):
    import os
    get_files = os.listdir(file_source)
    files = filter(lambda x: x.endswith(file_extension), get_files)
    for file in files:
        os.replace(os.path.join(file_source, file), os.path.join(file_destination, file))


def send_email(serv, login, passw, path, message, file_name):
    import smtplib
    import os
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    from platform import python_version

    with open('recipients.txt', 'r') as fl:
        dt_emails = fl.readlines()
        dt_emails = [line.rstrip() for line in dt_emails]

    server = serv
    user = login
    password = passw

    recipients = dt_emails
    sender = 'backend.distrilog@mail.ru'
    # subject = 'Заказ frenchpharmacy.ru'
    # text = 'Не пытайтесь ответить на это письмо. Оно сформировано и отправлено автоматически.'
    # text = text_message
    # html = '<html><head></head><body><p>' + text + '</p></body></html>'
    html = message

    filepath = path
    basename = os.path.basename(filepath)
    filesize = os.path.getsize(filepath)

    # msg = MIMEMultipart('alternative')
    msg = MIMEMultipart('mixed')
    msg['Subject'] = f"Frenchpharmacy {file_name[:-5]}"
    msg['From'] = 'robot frenchpharmacy <' + sender + '>'
    msg['To'] = ', '.join(recipients)
    msg['Reply-To'] = sender
    msg['Return-Path'] = sender
    msg['X-Mailer'] = 'Python/' + (python_version())

    #part_text = MIMEText(text, 'plain')
    part_html = MIMEText(html, 'html')
    part_file = MIMEBase('application', 'octet-stream; name="{}"'.format(basename))
    part_file.set_payload(open(filepath, "rb").read())
    part_file.add_header('Content-Description', basename)
    part_file.add_header('Content-Disposition', 'attachment; filename="{}"; size={}'.format(basename, filesize))
    encoders.encode_base64(part_file)

    #msg.attach(part_text)
    msg.attach(part_html)
    msg.attach(part_file)

    mail = smtplib.SMTP_SSL(server)
    mail.login(user, password)
    mail.sendmail(sender, recipients, msg.as_string())
    mail.quit()


def text_message(dataframe):
    import pandas
    text = pandas.DataFrame()
    list_index = [ind[0] for ind in dataframe.index]
    # list = [i + 1 for i in range(dataframe.shape[0])]
    # print(list)
    list_data_order = [ind[1].strftime('%d/%m/%y') for ind in dataframe.index]
    text['ИТОГО С НДС'] = dataframe['ИТОГО С НДС']
    text['ШТУК В ЗАКАЗЕ'] = dataframe['ШТ']
    text['ДАТА ЗАКАЗА'] = list_data_order
    #text['ЗАКАЗ'] = list_index
    text.index = list_index
    return text
